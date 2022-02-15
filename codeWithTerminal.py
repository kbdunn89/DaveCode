from openpyxl import load_workbook
from datetime import date, datetime
import calendar
import PySimpleGUI as sg

# For debug printout
debug = False

# Options (set to 'True' or 'False')
default_to_end_of_month = True
round_to_year = False
dateToday = False

# Don't change!
round_error = False

customDate = ''
customDateTF = False


class Account:
    def __init__(self, data):
        global round_to_year, round_error, dateToday
        self.entry = []
        self.startValue = data[0][1]
        self.startDate = get_date(data[0])
        tempDate = get_date(data[-1])
        checkTemp = datetime.now()
        checkDate = date(checkTemp.year, checkTemp.month, checkTemp.day)
        tempRound = tempDate.replace(month=self.startDate.month, day=self.startDate.day)
        checkDelta = checkDate - tempRound
        tempDelta = tempRound - tempDate

        if dateToday:
            self.endDate = checkDate
            round_error = False
        elif round_to_year and tempDelta.days >= 0 and checkDelta.days >= 0:
            self.endDate = tempRound
            round_error = False
        elif round_to_year and (tempDelta.days < 0 or checkDelta.days < 0):
            self.endDate = get_date(data[-1])
            round_error = True
        elif customDateTF:
            inputMonth = int(customDate[0:2])
            inputDay = int(customDate[3:5])
            inputYear = int(customDate[6:10])
            self.endDate = date(inputYear, inputMonth, inputDay)
        else:
            self.endDate = get_date(data[-1])
        delta = self.endDate - self.startDate
        self.duration = delta.days

    def add_entry(self, entry):
        self.entry.append(Entry(entry))

    def get_final_value(self):
        return acct.entry[-1].value

    def print_account(self):
        print("Start Date: ", self.startDate, "\tEnd Date: ", self.endDate, "\tDuration: ", self.duration, " days")
        for data in self.entry:
            data.print_entry()


class Entry:
    def __init__(self, data):
        length = len(data)
        if length <= 2:
            self.date = get_date(data)
            self.value = data[1]
            self.cashFlow = 0
        else:
            self.date = get_date(data)
            self.value = data[1]
            self.cashFlow = data[2]

    def print_entry(self):
        print("Date: ", self.date, "\tAccount Value: $", self.value, "\tCash Flow: $", self.cashFlow)


def get_date(rawData):
    global default_to_end_of_month
    tempData = rawData[0].split()
    year = int(tempData[1])
    month = get_month(tempData[0])
    if (len(rawData) > 2) and (rawData[3] < 32) and (rawData[3] >= 1):
        day = int(rawData[3])
    else:
        if default_to_end_of_month:
            day = calendar.monthrange(year, month)[1]
        else:
            day = 1
    return date(year, month, day)


def get_month(word):
    if word == 'January':
        return 1
    elif word == 'February':
        return 2
    elif word == 'March':
        return 3
    elif word == 'April':
        return 4
    elif word == 'May':
        return 5
    elif word == 'June':
        return 6
    elif word == 'July':
        return 7
    elif word == 'August':
        return 8
    elif word == 'September':
        return 9
    elif word == 'October':
        return 10
    elif word == 'November':
        return 11
    elif word == 'December':
        return 12
    else:
        return 0


def modifiedDietz(account):
    v0 = acct.startValue
    v1 = acct.get_final_value()
    cashFlowTotal = 0
    for entry in acct.entry:
        cashFlowTotal = cashFlowTotal + entry.cashFlow

    print("\n\n\n\nStart Value = $", v0, " End Value = $", v1, " Cash Flow Total = ", cashFlowTotal, " Duration: ",
          acct.duration)
    numerator = v1 - v0 - cashFlowTotal
    sumInDenominator = 0
    for entries in acct.entry:
        delta = entries.date - acct.startDate
        sumInDenominator = sumInDenominator + (((acct.duration - delta.days) / acct.duration) * entries.cashFlow)
    denominator = sumInDenominator + v0
    retVal = numerator / denominator
    # print("My calculated value: ", retVal * 100, "%")
    return retVal


def calculate(fname, monthEnd, todayEnd, customDateBool, custDate=''):
    global acct, default_to_end_of_month, round_error, dateToday, customDate, customDateTF
    dateToday = todayEnd
    customDate = custDate
    customDateTF = customDateBool
    default_to_end_of_month = monthEnd
    wb = load_workbook(filename=fname)
    ws = wb[wb.sheetnames[0]]
    temp = list()
    for row in ws.iter_rows(min_row=10, max_col=19):
        data = list()
        finished = False
        blankCell = 0
        cellNum = 1
        for cell in row:
            if cell.value and not (
                    cellNum == 2 or cellNum == 3 or cellNum == 4 or cellNum == 8 or cellNum == 11 or cellNum == 14 or cellNum == 17):
                data.append(cell.value)
            else:
                blankCell = blankCell + 1
            if blankCell >= 19 or cell.value == "YTD Return %":
                finished = True
                break
            cellNum = cellNum + 1
        if finished:
            break
        temp.append(data)
    wb.close()
    acct = Account(temp)
    for data in temp:
        while len(data) > 2:
            acct.add_entry(data)
            data.remove(data[2])
            data.remove(data[2])

    acct.entry.sort(key=lambda e: (e.date.year, e.date.month, e.date.day))
    acct.print_account()
    if debug:
        print("\n\n\nDebug: ", debug)
        print("Settings:\n\tDefault to end of month? ", default_to_end_of_month)
        print("\tAccounts in increments of a year? ", round_to_year)
        if round_error:
            print("\t\tCheck for date conflicts with data and end date selected.")
    return [modifiedDietz(acct), acct.startDate, acct.endDate, round_error]


def collapse(layout, key, visible):
    return sg.pin(sg.Column(layout, key=key, visible=visible))


def main():
    SYMBOL_UP = '▲'
    SYMBOL_DOWN = '▼'
    section = [[sg.Multiline(size=(35, 4), key='-OUTPUT-' + sg.WRITE_ONLY_KEY)], ]
    layout = [[sg.Radio('Default to the last of month', 'RADIO2', default=True, key='-month1-'),
               sg.Radio('Default to the 1st of the month', 'RADIO2', default=False, key='-month2-')],
              [sg.Radio('Set end date as today', 'RADIO1', default=False, key='-today-'),
               sg.Radio('Use final transaction as end date', 'RADIO1', default=True, key='-final-')],
              [sg.Radio('Use custom end date', 'RADIO1', default=False, key='-custom-'),
               sg.In(size=(12, 1), key='-custDate-'), sg.Text('mm/dd/yyyy')], [sg.Text('Document to open')],
              [sg.In(key='-fileInput-'), sg.FileBrowse(file_types=(("Excel Files", "*.xlsx"),))],
              [sg.Open(), sg.Cancel()], [sg.T(SYMBOL_UP, enable_events=True, k='-OPEN SEC-', text_color='white')],
              [collapse(section, '-SEC-', False)]]
    window = sg.Window('Modified Dietz Return Calculator', layout)

    opened1 = False

    while True:
        event, values = window.read()
        fname = values['-fileInput-']
        monthEnd = values['-month1-']
        todayEnd = values['-today-']
        customDateBool = values['-custom-']
        customDate = values['-custDate-']

        if event == sg.WIN_CLOSED or event == 'Cancel':
            break

        if event.startswith('-OPEN SEC-'):
            opened1 = not opened1
            window['-OPEN SEC-'].update(SYMBOL_DOWN if opened1 else SYMBOL_UP)
            window['-SEC-'].update(visible=opened1)

        if event == 'Open' and not fname:
            sg.popup("Oh whoops!", "No filename supplied")
            opened1 = True
            window['-OPEN SEC-'].update(SYMBOL_DOWN)
            window['-SEC-'].update(visible=opened1)

        elif event == 'Open' and fname:
            retVal = calculate(fname, monthEnd, todayEnd, customDateBool, customDate)
            sg.cprint('\n\nThe Modified Dietz Return for the account spanning ', retVal[1], ' to ', retVal[2],
                      ' is: {:.4f}%'.format(retVal[0] * 100), window=window, key='-OUTPUT-' + sg.WRITE_ONLY_KEY)
            opened1 = True
            window['-OPEN SEC-'].update(SYMBOL_DOWN)
            window['-SEC-'].update(visible=opened1)

    window.close()


if __name__ == "__main__":
    main()
